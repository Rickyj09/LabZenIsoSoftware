import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

from sqlalchemy import event, inspect
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from openpyxl import Workbook

from app import create_app
from app.extensions import db
from app.models.documentos import (
    DOCUMENTO_VIGOR_EXTERNO,
    DOCUMENTO_VIGOR_FORMATO,
    DOCUMENTO_VIGOR_INTERNO,
    Documento,
    DocumentoVersion,
    DocumentoVigorCatalogo,
)
from app.models.empresa import Empresa
from app.models.seguridad import Usuario
from app.services import document_vigor_import_service as import_module
from app.services.document_vigor_import_service import (
    DocumentVigorImportError,
    DocumentVigorImportService,
    DocumentVigorPasswordError,
    build_import_key,
    build_stable_identity,
    _parse_date,
)


class DocumentVigorImportTest(unittest.TestCase):
    def setUp(self):
        self.temp_directory = tempfile.TemporaryDirectory()
        self.app = create_app({
            "TESTING": True,
            "SQLALCHEMY_DATABASE_URI": "sqlite://",
            "SQLALCHEMY_ENGINE_OPTIONS": {},
            "WTF_CSRF_ENABLED": False,
        })
        self.context = self.app.app_context()
        self.context.push()
        db.create_all()
        self.next_id = 1000

        def assign_ids(session, _flush_context, _instances):
            for item in session.new:
                if isinstance(item, (DocumentoVigorCatalogo, Documento, DocumentoVersion, Empresa, Usuario)) and item.id is None:
                    item.id = self.next_id
                    self.next_id += 1

        self.assign_ids = assign_ids
        event.listen(Session, "before_flush", self.assign_ids)
        db.session.add_all([
            Empresa(id=101, nombre="Empresa uno"),
            Empresa(id=102, nombre="Empresa dos"),
            Usuario(
                id=201,
                empresa_id=101,
                nombre="Importador",
                apellido="Uno",
                email="importador@test",
                username="importador",
                password_hash="x",
                activo=True,
            ),
        ])
        db.session.commit()

    def tearDown(self):
        event.remove(Session, "before_flush", self.assign_ids)
        db.session.remove()
        db.drop_all()
        self.context.pop()
        self.temp_directory.cleanup()

    def path(self, name):
        return Path(self.temp_directory.name) / name

    def save_document_workbook(self, path, sheet_name, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.cell(1, 4, "LISTA DE DOCUMENTOS EN VIGOR")
        ws.cell(9, 1, "CÓDIGO")
        ws.cell(9, 2, "DIRECCIÓN")
        ws.cell(9, 3, "TÍTULO")
        ws.cell(9, 5, "USUARIOS DEL DOCUMENTO")
        ws.cell(9, 6, "LOCALIZACIÓN")
        ws.cell(9, 7, "TIPO DE DOCUMENTO")
        ws.cell(9, 8, "ALMACENAMIENTO")
        ws.cell(9, 9, "DESTINO FINAL")
        ws.cell(9, 10, "FECHA DE APROBACIÓN")
        ws.cell(9, 11, "REVISIÓN")
        ws.cell(11, 1, "DOCUMENTOS")
        start = 13
        for offset, row in enumerate(rows):
            for column, value in enumerate(row, start=1):
                ws.cell(start + offset, column, value)
        wb.save(path)

    def save_format_workbook(self, path, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "C"
        ws.cell(1, 3, "LISTA DE FORMATOS EN VIGOR")
        ws.cell(10, 1, "CÓDIGO")
        ws.cell(10, 2, "TÍTULO")
        ws.cell(10, 4, "REVISIÓN")
        ws.cell(10, 5, "FECHA DE VIGENCIA")
        ws.cell(10, 6, "CUSTODIO")
        ws.cell(10, 7, "ALMACENAMIENTO")
        ws.cell(11, 7, "ACCESO AL DOCUMENTO")
        ws.cell(11, 8, "LUGAR")
        ws.cell(11, 9, "PROTECCIÓN")
        ws.cell(10, 10, "FÍSICO/ELECTRÓNICO")
        ws.cell(10, 11, "DESTINO FINAL")
        start = 12
        for offset, row in enumerate(rows):
            for column, value in enumerate(row, start=1):
                ws.cell(start + offset, column, value)
        wb.save(path)

    def make_three_workbooks(self):
        internos = self.path("internos.xlsx")
        externos = self.path("externos.xlsx")
        formatos = self.path("formatos.xlsx")
        self.save_document_workbook(internos, "DOCUMENTOS INTERNOS", [
            ["PROCEDIMIENTOS GENERALES DE CALIDAD"],
            ["INT-001", "DIV", "Procedimiento interno", None, "DC", "SharePoint", "ELECTRÓNICO", "Carpeta PGC", "Gestión documental", datetime(2026, 1, 2), "00"],
            [None, None, None],
            [None, "ELABORADO", None, "REVISADO", None, "APROBADO"],
        ])
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "DOCUMENTOS INTERNOS "
        ws1.cell(9, 1, "CÓDIGO")
        ws1.cell(13, 1, "DUP-001")
        ws1.cell(13, 3, "No debe importarse")
        ws2 = wb.create_sheet("DOCUMENTOS EXTERNOS")
        ws2.cell(9, 1, "CÓDIGO")
        ws2.cell(9, 3, "TÍTULO")
        ws2.cell(11, 1, "DOCUMENTOS")
        ws2.cell(12, 1, "EXT-001")
        ws2.cell(12, 3, "Norma externa")
        ws2.cell(12, 5, "DC")
        ws2.cell(12, 6, "SharePoint")
        ws2.cell(12, 7, "ELECTRÓNICO")
        ws2.cell(12, 8, "Externos")
        ws2.cell(12, 9, "Gestión documental")
        ws2.cell(12, 10, datetime(2026, 2, 3))
        ws2.cell(12, 11, "00")
        wb.save(externos)
        self.save_format_workbook(formatos, [
            ["FORMATOS DE PROCEDIMIENTOS"],
            ["FOR-001", "Formato uno", None, "00", datetime(2026, 3, 4), "DC", "Usuarios", "SharePoint", "SGC", "ELECTRÓNICO", "Gestión documental"],
            [None, "ELABORADO", None, None, "REVISADO", None, None, None, "APROBADO"],
        ])
        return internos, externos, formatos

    def import_sample(self):
        internos, externos, formatos = self.make_three_workbooks()
        return DocumentVigorImportService().import_excel_files(
            empresa_id=101,
            internos_path=internos,
            externos_path=externos,
            formatos_path=formatos,
            usuario_id=201,
        )

    def test_model_indexes_constraints_and_tenant_columns_exist(self):
        columns = {column["name"] for column in inspect(db.engine).get_columns("documento_vigor_catalogo")}
        self.assertIn("empresa_id", columns)
        self.assertIn("tipo_listado", columns)
        self.assertIn("identidad_estable", columns)
        self.assertIn("ordinal_identidad", columns)
        self.assertIn("documento_version_id", columns)
        index_names = {index["name"] for index in inspect(db.engine).get_indexes("documento_vigor_catalogo")}
        self.assertIn("ix_documento_vigor_empresa_tipo_activo", index_names)
        self.assertIn("ix_documento_vigor_empresa_codigo", index_names)
        self.assertIn("ix_documento_vigor_empresa_tipo_identidad", index_names)

        first = DocumentoVigorCatalogo(
            id=301,
            empresa_id=101,
            tipo_listado=DOCUMENTO_VIGOR_INTERNO,
            clave_importacion="INTERNO|A|00",
            identidad_estable="CODIGO:A#1",
            ordinal_identidad=1,
            codigo="A",
            titulo="A",
            fuente_archivo="a.xlsx",
            fuente_hoja="DOCUMENTOS INTERNOS",
            fuente_fila=1,
            importado_en=datetime.now(),
        )
        duplicate = DocumentoVigorCatalogo(
            id=302,
            empresa_id=101,
            tipo_listado=DOCUMENTO_VIGOR_INTERNO,
            clave_importacion="INTERNO|A|00",
            identidad_estable="CODIGO:A#1",
            ordinal_identidad=1,
            codigo="A",
            titulo="A2",
            fuente_archivo="a.xlsx",
            fuente_hoja="DOCUMENTOS INTERNOS",
            fuente_fila=2,
            importado_en=datetime.now(),
        )
        db.session.add_all([first, duplicate])
        with self.assertRaises(IntegrityError):
            db.session.commit()
        db.session.rollback()

    def test_imports_three_types_sections_and_ignores_external_internal_sheet(self):
        result = self.import_sample()
        self.assertEqual(result.insertados, 3)
        rows = DocumentoVigorCatalogo.query.order_by(DocumentoVigorCatalogo.tipo_listado).all()
        self.assertEqual({row.tipo_listado for row in rows}, {DOCUMENTO_VIGOR_INTERNO, DOCUMENTO_VIGOR_EXTERNO, DOCUMENTO_VIGOR_FORMATO})
        self.assertEqual(DocumentoVigorCatalogo.query.filter_by(codigo="DUP-001").count(), 0)
        internal = DocumentoVigorCatalogo.query.filter_by(tipo_listado=DOCUMENTO_VIGOR_INTERNO).one()
        self.assertEqual(internal.seccion, "PROCEDIMIENTOS GENERALES DE CALIDAD")
        self.assertEqual(internal.lugar_almacenamiento, "SharePoint / Carpeta PGC")
        formatted = DocumentoVigorCatalogo.query.filter_by(tipo_listado=DOCUMENTO_VIGOR_FORMATO).one()
        self.assertEqual(formatted.seccion, "FORMATOS DE PROCEDIMIENTOS")
        self.assertEqual(formatted.fecha_vigencia, date(2026, 3, 4))

    def test_external_code_without_title_is_preserved(self):
        internos, externos, formatos = self.make_three_workbooks()
        wb = Workbook()
        ws = wb.active
        ws.title = "DOCUMENTOS INTERNOS "
        ws.cell(9, 1, "CÓDIGO")
        ws = wb.create_sheet("DOCUMENTOS EXTERNOS")
        ws.cell(9, 1, "CÓDIGO")
        ws.cell(9, 3, "TÍTULO")
        ws.cell(12, 1, "EXT-SIN-TITULO")
        ws.cell(12, 2, "DIV")
        ws.cell(12, 11, "00")
        wb.save(externos)
        DocumentVigorImportService().import_excel_files(
            empresa_id=101,
            internos_path=internos,
            externos_path=externos,
            formatos_path=formatos,
        )
        row = DocumentoVigorCatalogo.query.filter_by(codigo="EXT-SIN-TITULO").one()
        self.assertIsNone(row.titulo)

    def test_import_is_scoped_by_company(self):
        internos, externos, formatos = self.make_three_workbooks()
        service = DocumentVigorImportService()
        service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        service.import_excel_files(empresa_id=102, internos_path=internos, externos_path=externos, formatos_path=formatos)
        self.assertEqual(DocumentoVigorCatalogo.query.filter_by(empresa_id=101).count(), 3)
        self.assertEqual(DocumentoVigorCatalogo.query.filter_by(empresa_id=102).count(), 3)

    def test_import_is_idempotent_and_can_update_existing_record(self):
        internos, externos, formatos = self.make_three_workbooks()
        service = DocumentVigorImportService()
        first = service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        second = service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        self.assertEqual(first.insertados, 3)
        self.assertEqual(second.insertados, 0)
        self.assertEqual(second.actualizados, 0)
        self.assertEqual(DocumentoVigorCatalogo.query.filter_by(empresa_id=101).count(), 3)

        self.save_format_workbook(formatos, [
            ["FORMATOS DE PROCEDIMIENTOS"],
            ["FOR-001", "Formato uno", None, "00", datetime(2026, 3, 4), "AC", "Usuarios", "SharePoint", "SGC", "ELECTRÓNICO", "Gestión documental"],
        ])
        third = service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        self.assertEqual(third.actualizados, 1)
        self.assertEqual(DocumentoVigorCatalogo.query.filter_by(codigo="FOR-001").one().custodio, "AC")
        self.assertEqual(DocumentoVigorCatalogo.query.filter_by(empresa_id=101).count(), 3)

    def test_title_change_updates_without_duplicate(self):
        internos, externos, formatos = self.make_three_workbooks()
        service = DocumentVigorImportService()
        service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        self.save_format_workbook(formatos, [
            ["FORMATOS DE PROCEDIMIENTOS"],
            ["FOR-001", "Formato uno actualizado", None, "00", datetime(2026, 3, 4), "DC", "Usuarios", "SharePoint", "SGC", "ELECTRÃ“NICO", "GestiÃ³n documental"],
        ])
        result = service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        row = DocumentoVigorCatalogo.query.filter_by(codigo="FOR-001").one()
        self.assertEqual(result.actualizados, 1)
        self.assertEqual(row.titulo, "Formato uno actualizado")
        self.assertEqual(DocumentoVigorCatalogo.query.filter_by(codigo="FOR-001", activo=True).count(), 1)

    def test_revision_change_updates_without_duplicate(self):
        internos, externos, formatos = self.make_three_workbooks()
        service = DocumentVigorImportService()
        service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        self.save_format_workbook(formatos, [
            ["FORMATOS DE PROCEDIMIENTOS"],
            ["FOR-001", "Formato uno", None, "01", datetime(2026, 3, 4), "DC", "Usuarios", "SharePoint", "SGC", "ELECTRÃ“NICO", "GestiÃ³n documental"],
        ])
        result = service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        row = DocumentoVigorCatalogo.query.filter_by(codigo="FOR-001").one()
        self.assertEqual(result.actualizados, 1)
        self.assertEqual(row.revision, "01")
        self.assertEqual(DocumentoVigorCatalogo.query.filter_by(codigo="FOR-001", activo=True).count(), 1)

    def test_title_and_revision_change_updates_without_duplicate(self):
        internos, externos, formatos = self.make_three_workbooks()
        service = DocumentVigorImportService()
        service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        self.save_format_workbook(formatos, [
            ["FORMATOS DE PROCEDIMIENTOS"],
            ["FOR-001", "Formato uno v2", None, "02", datetime(2026, 3, 4), "DC", "Usuarios", "SharePoint", "SGC", "ELECTRÃ“NICO", "GestiÃ³n documental"],
        ])
        result = service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        row = DocumentoVigorCatalogo.query.filter_by(codigo="FOR-001").one()
        self.assertEqual(result.actualizados, 1)
        self.assertEqual(row.titulo, "Formato uno v2")
        self.assertEqual(row.revision, "02")
        self.assertEqual(DocumentoVigorCatalogo.query.filter_by(codigo="FOR-001", activo=True).count(), 1)

    def test_legitimate_repeated_code_uses_occurrence_identity(self):
        internos, externos, formatos = self.make_three_workbooks()
        self.save_format_workbook(formatos, [
            ["FORMATOS DE PROCEDIMIENTOS"],
            ["FOR-REP", "Formato repetido uno", None, "00", datetime(2026, 3, 4), "DC", "Usuarios", "SharePoint", "SGC", "ELECTRÃ“NICO", "GestiÃ³n documental"],
            ["FOR-REP", "Formato repetido dos", None, "00", datetime(2026, 3, 4), "DC", "Usuarios", "SharePoint", "SGC", "ELECTRÃ“NICO", "GestiÃ³n documental"],
        ])
        result = DocumentVigorImportService().import_excel_files(
            empresa_id=101,
            internos_path=internos,
            externos_path=externos,
            formatos_path=formatos,
        )
        rows = DocumentoVigorCatalogo.query.filter_by(codigo="FOR-REP").order_by(DocumentoVigorCatalogo.ordinal_identidad).all()
        self.assertEqual(result.insertados, 4)
        self.assertEqual(len(rows), 2)
        self.assertEqual([row.ordinal_identidad for row in rows], [1, 2])
        self.assertEqual([row.titulo for row in rows], ["Formato repetido uno", "Formato repetido dos"])

    def test_second_identical_import_omits_all_records_and_keeps_single_active_identity(self):
        internos, externos, formatos = self.make_three_workbooks()
        service = DocumentVigorImportService()
        service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        result = service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        self.assertEqual(result.insertados, 0)
        self.assertEqual(result.actualizados, 0)
        self.assertEqual(result.omitidos, 3)
        duplicates = (
            db.session.query(
                DocumentoVigorCatalogo.empresa_id,
                DocumentoVigorCatalogo.tipo_listado,
                DocumentoVigorCatalogo.identidad_estable,
                db.func.count(DocumentoVigorCatalogo.id),
            )
            .filter_by(activo=True)
            .group_by(
                DocumentoVigorCatalogo.empresa_id,
                DocumentoVigorCatalogo.tipo_listado,
                DocumentoVigorCatalogo.identidad_estable,
            )
            .having(db.func.count(DocumentoVigorCatalogo.id) > 1)
            .all()
        )
        self.assertEqual(duplicates, [])

    def test_rollback_when_fatal_error_occurs(self):
        internos, externos, formatos = self.make_three_workbooks()
        service = DocumentVigorImportService()
        original_import_sheet = service.import_sheet

        def fail_after_first(**kwargs):
            if kwargs["tipo_listado"] == DOCUMENTO_VIGOR_EXTERNO:
                raise DocumentVigorImportError("Falla controlada")
            return original_import_sheet(**kwargs)

        with patch.object(service, "import_sheet", side_effect=fail_after_first):
            with self.assertRaises(DocumentVigorImportError):
                service.import_excel_files(empresa_id=101, internos_path=internos, externos_path=externos, formatos_path=formatos)
        self.assertEqual(DocumentoVigorCatalogo.query.count(), 0)

    def test_password_errors_are_sanitized(self):
        with patch.object(import_module, "load_workbook", side_effect=Exception("SGC")):
            with self.assertRaises(DocumentVigorPasswordError) as ctx:
                import_module._open_workbook(self.path("missing.xlsx"), password="SGC")
        self.assertNotIn("SGC", str(ctx.exception))

    def test_missing_password_is_reported_without_secret(self):
        with patch.object(import_module, "load_workbook", side_effect=Exception("encrypted")):
            with self.assertRaises(DocumentVigorPasswordError) as ctx:
                import_module._open_workbook(self.path("missing.xlsx"))
        self.assertNotIn("SGC", str(ctx.exception))

    def test_optional_document_and_version_link(self):
        document = Documento(
            id=501,
            empresa_id=101,
            codigo="INT-001",
            titulo="Procedimiento interno",
            tipo_documento="PROCEDIMIENTO",
            estado="APROBADO",
            version_actual="00",
        )
        version = DocumentoVersion(
            id=601,
            empresa_id=101,
            documento_id=501,
            version="00",
            estado="APROBADO",
        )
        db.session.add_all([document, version])
        db.session.commit()
        self.import_sample()
        item = DocumentoVigorCatalogo.query.filter_by(codigo="INT-001").one()
        self.assertEqual(item.documento_id, 501)
        self.assertEqual(item.documento_version_id, 601)

    def test_import_key_uses_type_and_stable_identity(self):
        identity = build_stable_identity(" abc ", "Titulo", 1)
        self.assertEqual(identity, "CODIGO:ABC#1")
        self.assertEqual(len(build_import_key(DOCUMENTO_VIGOR_INTERNO, identity)), 64)
        self.assertEqual(
            build_import_key(DOCUMENTO_VIGOR_INTERNO, build_stable_identity(" abc ", "Titulo", 1)),
            build_import_key(DOCUMENTO_VIGOR_INTERNO, build_stable_identity("ABC", "Titulo nuevo", 1)),
        )
        self.assertNotEqual(
            build_import_key(DOCUMENTO_VIGOR_INTERNO, build_stable_identity("ABC", "Titulo", 1)),
            build_import_key(DOCUMENTO_VIGOR_EXTERNO, build_stable_identity("ABC", "Titulo", 1)),
        )
        self.assertNotEqual(
            build_import_key(DOCUMENTO_VIGOR_FORMATO, build_stable_identity("ABC", "Uno", 1)),
            build_import_key(DOCUMENTO_VIGOR_FORMATO, build_stable_identity("ABC", "Dos", 2)),
        )

    def test_unparseable_dates_are_treated_as_empty(self):
        self.assertIsNone(_parse_date("13/0/2025"))
        self.assertIsNone(_parse_date("N/A"))

    def test_unparseable_dates_generate_visible_warning(self):
        internos, externos, formatos = self.make_three_workbooks()
        self.save_format_workbook(formatos, [
            ["FORMATOS DE PROCEDIMIENTOS"],
            ["FOR-FECHA", "Formato con fecha invalida", None, "00", "13/0/2025", "DC", "Usuarios", "SharePoint", "SGC", "ELECTRÃ“NICO", "GestiÃ³n documental"],
        ])
        result = DocumentVigorImportService().import_excel_files(
            empresa_id=101,
            internos_path=internos,
            externos_path=externos,
            formatos_path=formatos,
        )
        row = DocumentoVigorCatalogo.query.filter_by(codigo="FOR-FECHA").one()
        self.assertIsNone(row.fecha_vigencia)
        self.assertEqual(len(result.advertencias), 1)
        self.assertIn("Fila 13: fecha invalida: 13/0/2025", result.advertencias[0])
