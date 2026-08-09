import io
import hashlib
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path

import msoffcrypto
from openpyxl import load_workbook

from app.extensions import db
from app.models.documentos import (
    DOCUMENTO_VIGOR_EXTERNO,
    DOCUMENTO_VIGOR_FORMATO,
    DOCUMENTO_VIGOR_INTERNO,
    Documento,
    DocumentoVersion,
    DocumentoVigorCatalogo,
)


class DocumentVigorImportError(ValueError):
    pass


class DocumentVigorPasswordError(DocumentVigorImportError):
    pass


@dataclass
class DocumentVigorImportSummary:
    tipo_listado: str
    fuente_archivo: str
    fuente_hoja: str
    insertados: int = 0
    actualizados: int = 0
    omitidos: int = 0
    errores: list[str] = field(default_factory=list)
    advertencias: list[str] = field(default_factory=list)

    @property
    def total_procesados(self):
        return self.insertados + self.actualizados


@dataclass
class DocumentVigorImportResult:
    summaries: list[DocumentVigorImportSummary] = field(default_factory=list)

    @property
    def insertados(self):
        return sum(item.insertados for item in self.summaries)

    @property
    def actualizados(self):
        return sum(item.actualizados for item in self.summaries)

    @property
    def omitidos(self):
        return sum(item.omitidos for item in self.summaries)

    @property
    def errores(self):
        errors = []
        for item in self.summaries:
            errors.extend(item.errores)
        return errors

    @property
    def advertencias(self):
        warnings = []
        for item in self.summaries:
            warnings.extend(item.advertencias)
        return warnings


_SPACE_RE = re.compile(r"\s+")
_SIGNATURE_LABELS = {"ELABORADO", "REVISADO", "APROBADO"}
_SIGNATURE_FOOTER_LABELS = {"NOMBRE", "CARGO", "FIRMA", "FECHA"}
_NON_SECTION_LABELS = {
    "LISTA DE DOCUMENTOS EN VIGOR",
    "LISTA DE FORMATOS EN VIGOR",
    "DOCUMENTOS",
    "DOC. INTERNOS",
    "DOC. EXTERNOS",
}


def _now():
    return datetime.now(timezone.utc)


def _clean_text(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).replace("_x0002_", "").strip()
    text = _SPACE_RE.sub(" ", text)
    return text or None


def _required_text(value):
    return _clean_text(value) or ""


def _clean_revision(value):
    text = _clean_text(value)
    if text is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return str(int(value)) if float(value).is_integer() else str(value)
    return text


def _parse_date(value):
    parsed, _warning = _parse_date_with_warning(value)
    return parsed


def _parse_date_with_warning(value):
    if value in (None, ""):
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    text = _clean_text(value)
    if not text:
        return None, None
    if text.upper() in {"N/A", "NA", "NO APLICA", "-"}:
        return None, None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date(), None
        except ValueError:
            continue
    return None, f"fecha invalida: {text}"


def _normalize_identity_text(value):
    return _required_text(value).upper()


def _identity_base(codigo, titulo):
    if codigo:
        return f"CODIGO:{_normalize_identity_text(codigo)}"
    normalized_title = _normalize_identity_text(titulo)
    title_hash = hashlib.sha256(normalized_title.encode("utf-8")).hexdigest()
    return f"SIN_CODIGO:{title_hash}"


def build_stable_identity(codigo, titulo, ordinal):
    return f"{_identity_base(codigo, titulo)}#{ordinal}"


def build_import_key(tipo_listado, identidad_estable):
    raw_key = f"{tipo_listado}|{identidad_estable}"
    return hashlib.sha256(raw_key.encode("utf-8")).hexdigest()


def _assign_stable_identities(records):
    counters = defaultdict(int)
    for record in records:
        base = _identity_base(record.get("codigo"), record.get("titulo"))
        counters[base] += 1
        ordinal = counters[base]
        identidad_estable = f"{base}#{ordinal}"
        record["ordinal_identidad"] = ordinal
        record["identidad_estable"] = identidad_estable
        record["clave_importacion"] = build_import_key(record["tipo_listado"], identidad_estable)
    return records


def _row_values(sheet, row_number):
    return [sheet.cell(row_number, column).value for column in range(1, sheet.max_column + 1)]


def _non_empty_texts(values):
    return [_clean_text(value) for value in values if _clean_text(value)]


def _is_signature_row(values):
    labels = {(_clean_text(value) or "").upper() for value in values}
    labels.discard("")
    return _SIGNATURE_LABELS.issubset(labels) or bool(labels) and labels.issubset(_SIGNATURE_FOOTER_LABELS)


def _is_header_row(values):
    labels = {(_clean_text(value) or "").upper() for value in values}
    return "CÓDIGO" in labels or "CODIGO" in labels


def _section_from_row(values):
    texts = _non_empty_texts(values)
    if len(texts) != 1:
        return None
    text = texts[0].strip()
    normalized = text.upper()
    if normalized in _NON_SECTION_LABELS:
        return None
    if any(label in normalized for label in ("PROCEDIMIENTO", "INSTRUCTIVO", "FORMATO", "DOCUMENTO", "OTRO")):
        return text
    return None


def _open_workbook(path, password=None):
    workbook_path = Path(path)
    try:
        return load_workbook(workbook_path, data_only=True)
    except Exception as original_error:
        if not password:
            raise DocumentVigorPasswordError(
                "No se pudo abrir un libro protegido; indica la contraseña."
            ) from original_error
        try:
            with workbook_path.open("rb") as source:
                office_file = msoffcrypto.OfficeFile(source)
                office_file.load_key(password=password)
                decrypted = io.BytesIO()
                office_file.decrypt(decrypted)
            decrypted.seek(0)
            return load_workbook(decrypted, data_only=True)
        except Exception as exc:
            raise DocumentVigorPasswordError(
                "No se pudo abrir el libro protegido; verifica la contraseña."
            ) from exc


def _parse_document_sheet(sheet, *, tipo_listado, fuente_archivo):
    records = []
    current_section = None
    for row_number in range(1, sheet.max_row + 1):
        values = _row_values(sheet, row_number)
        if _is_signature_row(values) or _is_header_row(values):
            continue
        section = _section_from_row(values)
        if section:
            current_section = section
            continue
        codigo = _clean_text(values[0] if len(values) > 0 else None)
        titulo = _clean_text(values[2] if len(values) > 2 else None)
        if not codigo and not titulo:
            continue
        if not codigo and (titulo or "").upper() in _NON_SECTION_LABELS:
            continue
        if not codigo and (titulo or "").upper() in _SIGNATURE_FOOTER_LABELS:
            continue
        if (codigo or "").upper() in _NON_SECTION_LABELS and not titulo:
            continue
        if codigo and not titulo and len(_non_empty_texts(values)) <= 1:
            continue
        date_value, date_warning = _parse_date_with_warning(values[9] if len(values) > 9 else None)
        location = _clean_text(values[5] if len(values) > 5 else None)
        storage = _clean_text(values[7] if len(values) > 7 else None)
        records.append({
            "tipo_listado": tipo_listado,
            "codigo": codigo,
            "titulo": titulo,
            "revision": _clean_revision(values[10] if len(values) > 10 else None),
            "fecha_vigencia": date_value,
            "custodio": None,
            "acceso_documento": _clean_text(values[4] if len(values) > 4 else None),
            "lugar_almacenamiento": " / ".join(item for item in (location, storage) if item) or None,
            "proteccion": None,
            "medio": _clean_text(values[6] if len(values) > 6 else None),
            "destino_final": _clean_text(values[8] if len(values) > 8 else None),
            "seccion": current_section or "DOCUMENTOS",
            "fuente_archivo": fuente_archivo,
            "fuente_hoja": sheet.title,
            "fuente_fila": row_number,
            "_advertencias": [date_warning] if date_warning else [],
        })
    return records


def _parse_format_sheet(sheet, *, fuente_archivo):
    records = []
    current_section = None
    for row_number in range(1, sheet.max_row + 1):
        values = _row_values(sheet, row_number)
        if _is_signature_row(values) or _is_header_row(values):
            continue
        section = _section_from_row(values)
        if section:
            current_section = section
            continue
        codigo = _clean_text(values[0] if len(values) > 0 else None)
        titulo = _clean_text(values[1] if len(values) > 1 else None)
        if not codigo and not titulo:
            continue
        if not codigo and (titulo or "").upper() in _NON_SECTION_LABELS:
            continue
        if not codigo and (titulo or "").upper() in _SIGNATURE_FOOTER_LABELS:
            continue
        if (codigo or "").upper() in _NON_SECTION_LABELS and not titulo:
            continue
        if codigo and not titulo and len(_non_empty_texts(values)) <= 1:
            continue
        date_value, date_warning = _parse_date_with_warning(values[4] if len(values) > 4 else None)
        records.append({
            "tipo_listado": DOCUMENTO_VIGOR_FORMATO,
            "codigo": codigo,
            "titulo": titulo,
            "revision": _clean_revision(values[3] if len(values) > 3 else None),
            "fecha_vigencia": date_value,
            "custodio": _clean_text(values[5] if len(values) > 5 else None),
            "acceso_documento": _clean_text(values[6] if len(values) > 6 else None),
            "lugar_almacenamiento": _clean_text(values[7] if len(values) > 7 else None),
            "proteccion": _clean_text(values[8] if len(values) > 8 else None),
            "medio": _clean_text(values[9] if len(values) > 9 else None),
            "destino_final": _clean_text(values[10] if len(values) > 10 else None),
            "seccion": current_section,
            "fuente_archivo": fuente_archivo,
            "fuente_hoja": sheet.title,
            "fuente_fila": row_number,
            "_advertencias": [date_warning] if date_warning else [],
        })
    return records


class DocumentVigorImportService:
    def __init__(self, session=None):
        self.session = session or db.session

    def import_excel_files(
        self,
        *,
        empresa_id,
        internos_path,
        externos_path,
        formatos_path,
        password=None,
        usuario_id=None,
        commit=True,
    ):
        result = DocumentVigorImportResult()
        try:
            result.summaries.append(self.import_sheet(
                empresa_id=empresa_id,
                tipo_listado=DOCUMENTO_VIGOR_INTERNO,
                path=internos_path,
                sheet_name="DOCUMENTOS INTERNOS",
                password=password,
                usuario_id=usuario_id,
            ))
            result.summaries.append(self.import_sheet(
                empresa_id=empresa_id,
                tipo_listado=DOCUMENTO_VIGOR_EXTERNO,
                path=externos_path,
                sheet_name="DOCUMENTOS EXTERNOS",
                password=password,
                usuario_id=usuario_id,
            ))
            result.summaries.append(self.import_sheet(
                empresa_id=empresa_id,
                tipo_listado=DOCUMENTO_VIGOR_FORMATO,
                path=formatos_path,
                sheet_name=None,
                password=None,
                usuario_id=usuario_id,
            ))
            if commit:
                self.session.commit()
        except Exception:
            if commit:
                self.session.rollback()
            raise
        return result

    def import_sheet(self, *, empresa_id, tipo_listado, path, sheet_name, password=None, usuario_id=None):
        workbook_path = Path(path)
        workbook = _open_workbook(workbook_path, password=password)
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise DocumentVigorImportError(f"No existe la hoja requerida: {sheet_name}.")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.worksheets[0]
        summary = DocumentVigorImportSummary(
            tipo_listado=tipo_listado,
            fuente_archivo=workbook_path.name,
            fuente_hoja=sheet.title,
        )
        if tipo_listado == DOCUMENTO_VIGOR_FORMATO:
            records = _parse_format_sheet(sheet, fuente_archivo=workbook_path.name)
        else:
            records = _parse_document_sheet(sheet, tipo_listado=tipo_listado, fuente_archivo=workbook_path.name)
        _assign_stable_identities(records)
        for record in records:
            for warning in record.pop("_advertencias", []):
                summary.advertencias.append(f"Fila {record.get('fuente_fila')}: {warning}")
            try:
                self._upsert_record(empresa_id=empresa_id, record=record, usuario_id=usuario_id, summary=summary)
            except DocumentVigorImportError as exc:
                summary.errores.append(f"Fila {record.get('fuente_fila')}: {exc}")
        return summary

    def _upsert_record(self, *, empresa_id, record, usuario_id, summary):
        if not record.get("codigo") and not record.get("titulo"):
            summary.omitidos += 1
            return
        documento_id, documento_version_id = self._resolve_document_link(
            empresa_id=empresa_id,
            codigo=record.get("codigo"),
            revision=record.get("revision"),
        )
        record["documento_id"] = documento_id
        record["documento_version_id"] = documento_version_id

        existing = (
            self.session.query(DocumentoVigorCatalogo)
            .filter_by(
                empresa_id=empresa_id,
                tipo_listado=record["tipo_listado"],
                clave_importacion=record["clave_importacion"],
            )
            .first()
        )
        now = _now()
        if not existing:
            self.session.add(DocumentoVigorCatalogo(
                empresa_id=empresa_id,
                activo=True,
                importado_por_id=usuario_id,
                importado_en=now,
                **record,
            ))
            summary.insertados += 1
            return

        changed = False
        for field_name, value in record.items():
            if getattr(existing, field_name) != value:
                setattr(existing, field_name, value)
                changed = True
        if not existing.activo:
            existing.activo = True
            changed = True
        if changed:
            existing.actualizado_por_id = usuario_id
            existing.actualizado_en = now
            summary.actualizados += 1
        else:
            summary.omitidos += 1

    def _resolve_document_link(self, *, empresa_id, codigo, revision=None):
        if not codigo:
            return None, None
        documento = (
            self.session.query(Documento)
            .filter_by(empresa_id=empresa_id, codigo=codigo)
            .first()
        )
        if not documento:
            return None, None
        version_query = self.session.query(DocumentoVersion).filter_by(
            empresa_id=empresa_id,
            documento_id=documento.id,
        )
        version_doc = None
        if revision:
            version_doc = version_query.filter_by(version=revision).first()
        if not version_doc and documento.version_vigente_id:
            version_doc = version_query.filter_by(id=documento.version_vigente_id).first()
        return documento.id, version_doc.id if version_doc else None
